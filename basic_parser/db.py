import sqlite3
from django.http import HttpResponse
from openpyxl import Workbook
from basic_parser.models import Profile, Skills
from django.conf import settings
import os
__author__ = 'Alex'

def clean():
    """method  to delete profiles without names"""

    profiles = Profile.objects.all()
    for profile in profiles:
        if profile.name == '' or profile.name is None:
            profile.delete()

def db_import(sm):
    """
    method to insert new profiles into DB or update
    if information is not up to date or a profile now is in 1st connection
    """
    for x in range(len(sm.urls)):
        name = sm.names[x]
        title = sm.titles[x]
        email = sm.emails[x]
        phone = sm.phones[x]
        im = sm.im[x]
        address = sm.address[x]
        connection = sm.connections[x]
        summary = sm.summaries[x]
        skills = str(sm.skills[x]).split(' | ')
        advice_to_connect = sm.advice_to_connect[x]
        url = sm.urls[x]

        if not Profile.objects.filter(url=url):
            profile = Profile(name=name, title=title, email=email, phone=phone,
                              im=im, address=address, summary=summary, advice_to_connect = advice_to_connect, url=url)
            profile.save()
            for skill in skills:
                if skill != '' and skill !='   ':
                        skills_obj = Skills(skill_name=skill.lower())
                        skills_obj.save()
                        profile.skills.add(skills_obj)

        else:
            if connection is not None:
                if connection == '1st':
                    profile = Profile.objects.get(url=url)
                    profile.email = email
                    profile.address = address
                    profile.phone = phone
                    profile.summary = summary
                    profile.im = im
                    profile.save()
                elif connection == '2nd':
                    profile = Profile.objects.get(url=url)
                    profile.name = name
                    profile.title = title
                    profile.address = address
                    profile.summary = profile.summary
                    for skill in skills:
                        if skill != '' and skill !='   ':
                            skills_obj = Skills(skill_name=skill.lower())
                            skills_obj.save()
                            profile.skills.add(skills_obj)
                    profile.save()
        clean()


def download_all(request):
    """this methods creates file with all profiles from db to be downloaded as xlsx file"""

    full_file_name = os.path.join(settings.MEDIA_ROOT, "all_profiles.xlsx")
    profiles = Profile.objects.all()
    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    for x in range(len(profiles)):
        ws.cell(column=1, row=x+1, value=profiles[x].name)
        ws.cell(column=2, row=x+1, value=profiles[x].title)
        ws.cell(column=3, row=x+1, value=(''.join([x for x in (str(profiles[x].email),str(profiles[x].im), str(profiles[x].phone), str(profiles[x].advice_to_connect)) if x !='None'])).replace('[', '').replace(']', ''))
        ws.cell(column=4, row=x+1, value=profiles[x].summary)
        ws.cell(column=5, row=x+1, value=(" | ".join([skill.skill_name for skill in Skills.objects.filter(profile=profiles[x])])))
        ws.cell(column=6, row=x+1, value=profiles[x].url)

    wb.save(filename = full_file_name)
    file = open(full_file_name, 'rb')
    response = HttpResponse(file, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=all_profiles.xls'
    return  response
