import os
import time
import xlrd
from django.conf import settings
from openpyxl import Workbook
from grab import Grab
from basic_parser.models import Profile


def links_list(request):
    """method creates links to be followed and parsed from uploaded files"""

    file = request.FILES.get('files')
    full_path = os.path.join(settings.BASE_DIR,'media/') + str(file)
    with open(full_path, 'wb+') as destinaton:
        for chunck in file.chunks():
            destinaton.write(chunck)

    sheet_index = request.POST.get('sheet number')
    column_index = request.POST.get('column number')
    p_start = request.POST.get('p-start')
    p_end = request.POST.get('p-end')

    wb = xlrd.open_workbook(filename=full_path, encoding_override='utf-8')
    worksheet = wb.sheet_by_index(int(sheet_index)-1)
    table = []
    for cell in worksheet.col(int(column_index)-1):
        if cell.value not in table:
            table.append(str(cell.value))
    return table[int(p_start)-1:int(p_end)+1]

def make_new_link(grab, url):
    """method to form new link from response.head['Location']"""
    try:
        grab.go(url)
        head = str(grab.response.head).split('\\r\\')

        new_link = ''
        for x in head:
            if 'https' in x:
                new_link = x.replace('https://ua.', 'https://www.').replace(" ", "").replace('nLocation:', '').strip(" ").replace('https://pl.', 'https://www.')
    except:
        new_link = ""
    return new_link


def xls_export(sm, name):
    """method to export to xls file profiles from current Spider"""

    dest_filename = os.path.join(settings.MEDIA_ROOT, '{0}: {1}.xls'.format(str(name), time.strftime("%H:%M:%S")))
    wb = Workbook()
    ws3 = wb.get_sheet_by_name('Sheet')
    for x in range(len(sm.names)):
        ws3.cell(column=1, row=x+1, value=sm.names[x])
        ws3.cell(column=2, row=x+1, value=sm.titles[x])
        ws3.cell(column=3, row=x+1, value=sm.connections[x])
        ws3.cell(column=4, row=x+1, value=str(sm.emails[x] + '\n' + sm.phones[x] + '\n' + sm.advice_to_connect[x]).strip())
        ws3.cell(column=5, row=x+1, value=sm.im[x])
        ws3.cell(column=6, row=x+1, value=sm.summaries[x])
        ws3.cell(column=7, row=x+1, value=sm.skills[x])
        ws3.cell(column=8, row=x+1, value=sm.urls[x])
    wb.save(filename=dest_filename)
    return dest_filename




def spider_login(request):
    """method to login into linkedin"""

    grab = Grab()
    login = request.POST.get('login')
    password = request.POST.get('password')
    grab.go("https://www.linkedin.com/uas/login")
    grab.doc.set_input('session_key', login)
    grab.doc.set_input('session_password', password)
    grab.doc.submit()
    return grab


def spider_xpaths():
    """method to create all requires fields that later will be parced"""

    x_path_elements = ['//*[@class="full-name"]',
                        '//*[@id="headline"]/p',
                        "//*[@class='fp-degree-icon']",
                        '//*[@id="email-view"]/ul/li/a', '//*[@id="phone-view"]',
                        '//*[@id="summary-item-view"]/div/p',
                        '//*[@class="endorse-item-name-text"]',
                        '//*[@id="im"]',
                        '//*[@id="location"]/dl/dd[1]/span/a',
                        '//*[@id="contact-comments-view"]']
    return x_path_elements


def spider_go_first(grab, url, x_path_elements, names, titles, connections, emails, phones, summaries, skills, urls, im, address, advice_to_connect,):
    """method to parse a profile if it is a 1st connection profile"""

    if grab.doc.select(x_path_elements[0]): names.append(grab.doc.select(x_path_elements[0]).text())
    else: names.append("")

    if grab.doc.select(x_path_elements[1]): titles.append(grab.doc.select(x_path_elements[1]).text())
    else: titles.append("")

    if grab.doc.select(x_path_elements[2]): connections.append(grab.doc.select(x_path_elements[2]).text())
    else: connections.append("")

    if grab.doc.select(x_path_elements[3]): emails.append(grab.doc.select(x_path_elements[3]).text())
    else: emails.append("")

    if grab.doc.select(x_path_elements[4]): phones.append(grab.doc.select(x_path_elements[4]).text())
    else: phones.append("")

    if grab.doc.select(x_path_elements[5]): summaries.append(grab.doc.select(x_path_elements[5]).text())
    else: summaries.append("")

    if grab.doc.select(x_path_elements[6]): skills.append( " | ".join([elem_in.text() for elem_in in grab.doc.select(x_path_elements[6])]))
    else: skills.append("")

    urls.append(url)

    if grab.doc.select(x_path_elements[7]): im.append(grab.doc.select(x_path_elements[7]).text())
    else: im.append("")

    if grab.doc.select(x_path_elements[8]): address.append(grab.doc.select(x_path_elements[8]).text())
    else: address.append("")

    if grab.doc.select(x_path_elements[9]): advice_to_connect.append(grab.doc.select(x_path_elements[9]).text())
    else: advice_to_connect.append("")



def spider_go_second(grab, url, x_path_elements, names, titles, connections, emails, phones, summaries, skills, urls, im, address, advice_to_connect,):
    """
    method to check if the profile is already in the DB,
    and if it is - take info from it,
    if not - parse it by calling above method
    """
    try:
        profile = Profile.objects.get(url=url)
        names.append(profile.name)
        connections.append(grab.doc.select(x_path_elements[2]).text())
        titles.append(profile.title)
        emails.append(profile.email)
        phones.append(profile.phone)
        summaries.append(profile.summary)
        if grab.doc.select(x_path_elements[6]): skills.append( " | ".join([elem_in.text() for elem_in in grab.doc.select(x_path_elements[6])]))
        else: skills.append("")
        im.append(profile.im)
        address.append(profile.address)
        advice_to_connect.append(profile.advice_to_connect)
        urls.append(url)
    except:
        spider_go_first(grab, url, x_path_elements, names, titles, connections, emails, phones, summaries, skills, urls, im, address, advice_to_connect)

